import wx
import os
import openpyxl
from openpyxl import Workbook

# ===== Config & Colors
DATA_DIR = os.path.join("data", "todolist")
HEADER = ["Task", "Status", "Note"]  # Status: 0 = aktif, 1 = selesai

YELLOW = wx.Colour(255, 255, 102)      # task aktif
GREY_DONE = wx.Colour(136, 136, 136)   # task selesai
GREY_NOTE = wx.Colour(180, 180, 180)   # note
BG_DARK = wx.Colour(30, 30, 30)
BG_PANEL = wx.Colour(40, 40, 40)
FG_TEXT = wx.Colour(230, 230, 230)


# ===== Helpers
def ensure_data_dir():
    os.makedirs(DATA_DIR, exist_ok=True)


def create_new_todo_excel(path, first_sheet_name="Default"):
    wb = Workbook()
    ws = wb.active
    ws.title = first_sheet_name
    ws.append(HEADER)
    wb.save(path)


def calc_todo_progress(path) -> int:
    """Hitung persentase selesai dari semua sheet dalam file Excel."""
    try:
        wb = openpyxl.load_workbook(path)
    except Exception:
        return 0
    total = 0
    done = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, values_only=True):
            task, status, note = row[:3]
            if task and str(task).strip() != "":
                total += 1
                if status in (1, "1", True):
                    done += 1
    if total == 0:
        return 0
    return int((done / total) * 100)


def parse_display_to_name(display_text: str) -> str:
    """Sidebar menampilkan 'Nama (80%)' → ambil 'Nama' saja."""
    if " (" in display_text and display_text.endswith(")"):
        return display_text.rsplit(" (", 1)[0]
    return display_text


# ===== Dialogs
class ItemDialog(wx.Dialog):
    """Dialog tambah/edit task: task + note dalam 1 submit."""
    def __init__(self, parent, title="Item", task="", note=""):
        super().__init__(parent, title=title, size=(430, 260))
        pnl = wx.Panel(self)
        pnl.SetBackgroundColour(BG_DARK)

        v = wx.BoxSizer(wx.VERTICAL)

        t1 = wx.StaticText(pnl, label="Nama Task")
        t1.SetForegroundColour(FG_TEXT)
        self.txt_task = wx.TextCtrl(pnl, value=task)
        self.txt_task.SetBackgroundColour(BG_PANEL)
        self.txt_task.SetForegroundColour(FG_TEXT)

        t2 = wx.StaticText(pnl, label="Keterangan (opsional)")
        t2.SetForegroundColour(FG_TEXT)
        self.txt_note = wx.TextCtrl(pnl, value=note, style=wx.TE_MULTILINE)
        self.txt_note.SetBackgroundColour(BG_PANEL)
        self.txt_note.SetForegroundColour(FG_TEXT)

        btns = wx.StdDialogButtonSizer()
        ok_btn = wx.Button(pnl, wx.ID_OK)
        cancel_btn = wx.Button(pnl, wx.ID_CANCEL)
        btns.AddButton(ok_btn)
        btns.AddButton(cancel_btn)
        btns.Realize()

        v.Add(t1, 0, wx.ALL, 8)
        v.Add(self.txt_task, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 8)
        v.Add(t2, 0, wx.LEFT | wx.RIGHT | wx.TOP, 8)
        v.Add(self.txt_note, 1, wx.EXPAND | wx.ALL, 8)
        v.Add(btns, 0, wx.EXPAND | wx.ALL, 8)

        pnl.SetSizer(v)
        self.Centre()

    def get_values(self):
        return self.txt_task.GetValue().strip(), self.txt_note.GetValue().strip()


class SectionDialog(wx.Dialog):
    """Dialog tambah/rename section (sheet)."""
    def __init__(self, parent, title="Section", name=""):
        super().__init__(parent, title=title, size=(360, 160))
        pnl = wx.Panel(self)
        pnl.SetBackgroundColour(BG_DARK)
        v = wx.BoxSizer(wx.VERTICAL)

        t1 = wx.StaticText(pnl, label="Nama Section")
        t1.SetForegroundColour(FG_TEXT)
        self.txt = wx.TextCtrl(pnl, value=name)
        self.txt.SetBackgroundColour(BG_PANEL)
        self.txt.SetForegroundColour(FG_TEXT)

        btns = wx.StdDialogButtonSizer()
        ok_btn = wx.Button(pnl, wx.ID_OK)
        cancel_btn = wx.Button(pnl, wx.ID_CANCEL)
        btns.AddButton(ok_btn)
        btns.AddButton(cancel_btn)
        btns.Realize()

        v.Add(t1, 0, wx.ALL, 10)
        v.Add(self.txt, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 10)
        v.Add(btns, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 10)

        pnl.SetSizer(v)
        self.Centre()

    def get_value(self):
        return self.txt.GetValue().strip()


class RenameTodoDialog(wx.Dialog):
    """Dialog create/rename todo (nama file tanpa .xlsx)."""
    def __init__(self, parent, old_name=""):
        super().__init__(parent, title="Nama Todo", size=(360, 160))
        pnl = wx.Panel(self)
        pnl.SetBackgroundColour(BG_DARK)
        v = wx.BoxSizer(wx.VERTICAL)

        t1 = wx.StaticText(pnl, label="Nama Todo")
        t1.SetForegroundColour(FG_TEXT)
        self.txt = wx.TextCtrl(pnl, value=old_name)
        self.txt.SetBackgroundColour(BG_PANEL)
        self.txt.SetForegroundColour(FG_TEXT)

        btns = wx.StdDialogButtonSizer()
        ok_btn = wx.Button(pnl, wx.ID_OK)
        cancel_btn = wx.Button(pnl, wx.ID_CANCEL)
        btns.AddButton(ok_btn)
        btns.AddButton(cancel_btn)
        btns.Realize()

        v.Add(t1, 0, wx.ALL, 10)
        v.Add(self.txt, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 10)
        v.Add(btns, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 10)

        pnl.SetSizer(v)
        self.Centre()

    def get_value(self):
        return self.txt.GetValue().strip()


# ===== Main App
class TodoApp(wx.Frame):
    def __init__(self):
        super().__init__(None, title="Excel Todo Manager", size=(1100, 700))
        ensure_data_dir()
        self.SetBackgroundColour(BG_DARK)

        self.file_map = {}       # display_name (tanpa ekstensi) -> full path
        self.current_file = None # full path saat ini
        self.current_name = ""   # display name (tanpa .xlsx)
        self.current_sheet = None

        root = wx.BoxSizer(wx.HORIZONTAL)

        # ==== Sidebar
        side_panel = wx.Panel(self)
        side_panel.SetBackgroundColour(BG_PANEL)
        side_sizer = wx.BoxSizer(wx.VERTICAL)

        lbl = wx.StaticText(side_panel, label="Daftar Todo")
        lbl.SetForegroundColour(FG_TEXT)
        f = lbl.GetFont(); f.MakeBold(); lbl.SetFont(f)

        self.todo_list = wx.ListBox(side_panel, size=(300, -1))
        self.todo_list.Bind(wx.EVT_LISTBOX, self.on_select_todo)
        self.todo_list.SetBackgroundColour(BG_DARK)
        self.todo_list.SetForegroundColour(FG_TEXT)

        btn_new = wx.Button(side_panel, label="Buat Todo Baru")
        btn_ren = wx.Button(side_panel, label="Rename Todo")
        btn_del = wx.Button(side_panel, label="Hapus Todo")
        btn_ref = wx.Button(side_panel, label="Refresh")
        for b in (btn_new, btn_ren, btn_del, btn_ref):
            b.SetBackgroundColour(wx.Colour(0, 122, 204))
            b.SetForegroundColour(wx.WHITE)
        btn_new.Bind(wx.EVT_BUTTON, self.create_todo)
        btn_ren.Bind(wx.EVT_BUTTON, self.rename_todo)
        btn_del.Bind(wx.EVT_BUTTON, self.delete_todo)
        btn_ref.Bind(wx.EVT_BUTTON, lambda e: self.load_todo_files(preserve=self.current_name))

        side_sizer.Add(lbl, 0, wx.ALL, 8)
        side_sizer.Add(self.todo_list, 1, wx.EXPAND | wx.ALL, 8)
        side_sizer.Add(btn_new, 0, wx.EXPAND | wx.ALL, 6)
        side_sizer.Add(btn_ren, 0, wx.EXPAND | wx.ALL, 6)
        side_sizer.Add(btn_del, 0, wx.EXPAND | wx.ALL, 6)
        side_sizer.Add(btn_ref, 0, wx.EXPAND | wx.ALL, 6)
        side_panel.SetSizer(side_sizer)

        root.Add(side_panel, 0, wx.EXPAND)

        # ==== Main area
        main_panel = wx.Panel(self)
        main_panel.SetBackgroundColour(BG_DARK)
        main_sizer = wx.BoxSizer(wx.VERTICAL)

        self.lbl_title = wx.StaticText(main_panel, label="(Belum memilih Todo)")
        self.lbl_title.SetForegroundColour(FG_TEXT)
        ft = self.lbl_title.GetFont(); ft.MakeBold(); ft.SetPointSize(ft.GetPointSize() + 2)
        self.lbl_title.SetFont(ft)
        main_sizer.Add(self.lbl_title, 0, wx.ALL, 10)

        # Section controls
        sheet_bar = wx.BoxSizer(wx.HORIZONTAL)
        sheet_bar.Add(wx.StaticText(main_panel, label="Section:"), 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 6)

        self.sheet_list = wx.Choice(main_panel, choices=[])
        self.sheet_list.Bind(wx.EVT_CHOICE, self.on_select_sheet)
        self.sheet_list.SetBackgroundColour(BG_PANEL)
        self.sheet_list.SetForegroundColour(FG_TEXT)
        sheet_bar.Add(self.sheet_list, 0, wx.RIGHT, 12)

        btn_add_sec = wx.Button(main_panel, label="Tambah Section")
        btn_ren_sec = wx.Button(main_panel, label="Rename Section")
        btn_del_sec = wx.Button(main_panel, label="Hapus Section")
        for b in (btn_add_sec, btn_ren_sec, btn_del_sec):
            b.SetBackgroundColour(wx.Colour(85, 85, 85))
            b.SetForegroundColour(wx.WHITE)
        btn_add_sec.Bind(wx.EVT_BUTTON, self.add_section)
        btn_ren_sec.Bind(wx.EVT_BUTTON, self.rename_section)
        btn_del_sec.Bind(wx.EVT_BUTTON, self.delete_section)

        sheet_bar.Add(btn_add_sec, 0, wx.RIGHT, 6)
        sheet_bar.Add(btn_ren_sec, 0, wx.RIGHT, 6)
        sheet_bar.Add(btn_del_sec, 0)
        main_sizer.Add(sheet_bar, 0, wx.ALL, 10)

        # Task bar
        task_bar = wx.BoxSizer(wx.HORIZONTAL)
        btn_add_item = wx.Button(main_panel, label="Tambah Task")
        btn_add_item.SetBackgroundColour(wx.Colour(0, 122, 204))
        btn_add_item.SetForegroundColour(wx.WHITE)
        btn_add_item.Bind(wx.EVT_BUTTON, self.add_item)
        task_bar.Add(btn_add_item, 0, wx.RIGHT, 6)
        main_sizer.Add(task_bar, 0, wx.LEFT | wx.RIGHT | wx.BOTTOM, 10)

        # Scrolled task area
        self.scroll = wx.ScrolledWindow(main_panel, style=wx.VSCROLL)
        self.scroll.SetScrollRate(0, 14)
        self.scroll.SetBackgroundColour(BG_DARK)
        self.task_area = wx.Panel(self.scroll)
        self.task_area.SetBackgroundColour(BG_DARK)
        self.task_sizer = wx.BoxSizer(wx.VERTICAL)
        self.task_area.SetSizer(self.task_sizer)
        sv = wx.BoxSizer(wx.VERTICAL)
        sv.Add(self.task_area, 1, wx.EXPAND | wx.ALL, 6)
        self.scroll.SetSizer(sv)
        main_sizer.Add(self.scroll, 1, wx.EXPAND | wx.ALL, 10)

        main_panel.SetSizer(main_sizer)
        root.Add(main_panel, 1, wx.EXPAND)

        self.SetSizer(root)
        self.Centre()
        self.load_todo_files()
        self.Show()

    # ===== Sidebar & Files =====
    def load_todo_files(self, preserve: str | None = None):
        """Muat semua file .xlsx → tampilkan 'Nama (XX%)'. Optionally preserve selection by name."""
        self.file_map.clear()
        displays = []
        for f in sorted(os.listdir(DATA_DIR)):
            if f.lower().endswith(".xlsx"):
                name = os.path.splitext(f)[0]
                path = os.path.join(DATA_DIR, f)
                self.file_map[name] = path
                pct = calc_todo_progress(path)
                displays.append(f"{name} ({pct}%)")
        self.todo_list.Set(displays)

        # restore selection if possible
        if preserve:
            for i, txt in enumerate(self.todo_list.GetStrings()):
                if parse_display_to_name(txt) == preserve:
                    self.todo_list.SetSelection(i)
                    break

    def get_selected_todo_name(self):
        sel = self.todo_list.GetSelection()
        if sel == wx.NOT_FOUND:
            return None
        return parse_display_to_name(self.todo_list.GetString(sel))

    def create_todo(self, event):
        dlg = RenameTodoDialog(self, old_name="")
        if dlg.ShowModal() == wx.ID_OK:
            name = dlg.get_value()
            if not name:
                wx.MessageBox("Nama tidak boleh kosong.", "Error", wx.OK | wx.ICON_ERROR)
                dlg.Destroy()
                return
            path = os.path.join(DATA_DIR, f"{name}.xlsx")
            if os.path.exists(path):
                wx.MessageBox("Nama sudah digunakan.", "Error", wx.OK | wx.ICON_ERROR)
                dlg.Destroy()
                return
            create_new_todo_excel(path, first_sheet_name="Default")
            self.load_todo_files(preserve=name)
        dlg.Destroy()

    def rename_todo(self, event):
        name = self.get_selected_todo_name()
        if not name:
            wx.MessageBox("Pilih todo dahulu.", "Info")
            return
        dlg = RenameTodoDialog(self, old_name=name)
        if dlg.ShowModal() == wx.ID_OK:
            new_name = dlg.get_value()
            if not new_name:
                wx.MessageBox("Nama tidak boleh kosong.", "Error")
                dlg.Destroy()
                return
            old_path = self.file_map[name]
            new_path = os.path.join(DATA_DIR, f"{new_name}.xlsx")
            if os.path.exists(new_path):
                wx.MessageBox("Nama sudah digunakan.", "Error")
                dlg.Destroy()
                return
            os.rename(old_path, new_path)
            # update state jika sedang dibuka
            if self.current_file == old_path:
                self.current_file = new_path
                self.current_name = new_name
                self.lbl_title.SetLabel(f"# {self.current_name}")
            self.load_todo_files(preserve=new_name)
        dlg.Destroy()

    def delete_todo(self, event):
        name = self.get_selected_todo_name()
        if not name:
            wx.MessageBox("Pilih todo dahulu.", "Info")
            return
        path = self.file_map.get(name)
        if wx.MessageBox(f"Hapus todo '{name}'?", "Konfirmasi",
                         wx.YES_NO | wx.ICON_WARNING) == wx.YES:
            try:
                os.remove(path)
            except Exception as e:
                wx.MessageBox(str(e), "Error")
                return
            if self.current_name == name:
                self.current_file = None
                self.current_name = ""
                self.current_sheet = None
                self.lbl_title.SetLabel("(Belum memilih Todo)")
                self.sheet_list.Set([])
                self.clear_tasks()
            self.load_todo_files()

    def on_select_todo(self, event):
        name = self.get_selected_todo_name()
        if not name:
            return
        path = self.file_map[name]
        self.current_file = path
        self.current_name = name
        self.lbl_title.SetLabel(f"# {self.current_name}")
        self.load_sheets()

    # ===== Sections (Sheets) =====
    def on_select_sheet(self, event):
        self.current_sheet = self.sheet_list.GetStringSelection()
        self.show_tasks()

    def load_sheets(self):
        try:
            wb = openpyxl.load_workbook(self.current_file)
        except Exception as e:
            wx.MessageBox(str(e), "Error")
            return
        names = wb.sheetnames
        self.sheet_list.Set(names)
        if not names:
            self.current_sheet = None
            self.clear_tasks()
            return
        # pilih yang sebelumnya, kalau tidak ada pilih pertama
        if self.current_sheet in names:
            idx = names.index(self.current_sheet)
        else:
            idx = 0
            self.current_sheet = names[0]
        self.sheet_list.SetSelection(idx)
        self.show_tasks()

    def add_section(self, event):
        if not self.current_file:
            return
        dlg = SectionDialog(self, title="Tambah Section")
        if dlg.ShowModal() == wx.ID_OK:
            name = dlg.get_value()
            if not name:
                dlg.Destroy(); return
            wb = openpyxl.load_workbook(self.current_file)
            if name in wb.sheetnames:
                wx.MessageBox("Section sudah ada.", "Error")
                dlg.Destroy(); return
            ws = wb.create_sheet(name)
            ws.append(HEADER)
            wb.save(self.current_file)
            self.current_sheet = name
            self.load_sheets()
        dlg.Destroy()

    def rename_section(self, event):
        if not self.current_file or not self.current_sheet:
            return
        dlg = SectionDialog(self, title="Rename Section", name=self.current_sheet)
        if dlg.ShowModal() == wx.ID_OK:
            new_name = dlg.get_value()
            if not new_name:
                dlg.Destroy(); return
            wb = openpyxl.load_workbook(self.current_file)
            if new_name in wb.sheetnames:
                wx.MessageBox("Section sudah ada.", "Error")
                dlg.Destroy(); return
            ws = wb[self.current_sheet]
            ws.title = new_name
            wb.save(self.current_file)
            self.current_sheet = new_name
            self.load_sheets()
        dlg.Destroy()

    def delete_section(self, event):
        if not self.current_file or not self.current_sheet:
            return
        wb = openpyxl.load_workbook(self.current_file)
        if len(wb.sheetnames) <= 1:
            wx.MessageBox("Minimal harus ada 1 section.", "Info")
            return
        if wx.MessageBox(f"Hapus section '{self.current_sheet}'?", "Konfirmasi",
                         wx.YES_NO | wx.ICON_WARNING) != wx.YES:
            return
        ws = wb[self.current_sheet]
        wb.remove(ws)
        wb.save(self.current_file)
        self.current_sheet = None
        self.load_sheets()

    # ===== Tasks (Items) =====
    def clear_tasks(self):
        for c in self.task_area.GetChildren():
            c.Destroy()
        self.task_sizer.Layout()
        self.scroll.Layout()

    def show_tasks(self):
        self.clear_tasks()
        if not self.current_file or not self.current_sheet:
            return
        wb = openpyxl.load_workbook(self.current_file)
        ws = wb[self.current_sheet]
        if ws.max_row == 0:
            ws.append(HEADER)
            wb.save(self.current_file)

        for row in ws.iter_rows(min_row=2, values_only=True):
            task, status, note = row[:3]
            if not task:
                continue
            status = int(status) if status else 0
            note = note or ""

            row_panel = wx.Panel(self.task_area)
            row_panel.SetBackgroundColour(BG_PANEL)
            vs = wx.BoxSizer(wx.VERTICAL)

            top = wx.BoxSizer(wx.HORIZONTAL)
            cb = wx.CheckBox(row_panel, label=str(task))
            cb.SetValue(bool(status))
            cb.SetForegroundColour(GREY_DONE if status else YELLOW)
            cb.Bind(wx.EVT_CHECKBOX, lambda e, t=str(task): self.toggle_task(e, t))

            btn_edit = wx.Button(row_panel, label="Edit", size=(72, 26))
            btn_del = wx.Button(row_panel, label="Hapus", size=(72, 26))
            btn_edit.SetBackgroundColour(wx.Colour(85, 85, 85)); btn_edit.SetForegroundColour(wx.WHITE)
            btn_del.SetBackgroundColour(wx.Colour(170, 0, 0));   btn_del.SetForegroundColour(wx.WHITE)
            btn_edit.Bind(wx.EVT_BUTTON, lambda e, t=str(task): self.edit_item(t))
            btn_del.Bind(wx.EVT_BUTTON,  lambda e, t=str(task): self.delete_item(t))

            top.Add(cb, 1, wx.ALL | wx.ALIGN_CENTER_VERTICAL, 6)
            top.Add(btn_edit, 0, wx.ALL, 6)
            top.Add(btn_del, 0, wx.ALL, 6)

            vs.Add(top, 0, wx.EXPAND)
            if note:
                txt = wx.StaticText(row_panel, label=str(note))
                txt.SetForegroundColour(GREY_NOTE)
                vs.Add(txt, 0, wx.LEFT | wx.RIGHT | wx.BOTTOM, 10)

            row_panel.SetSizer(vs)
            self.task_sizer.Add(row_panel, 0, wx.EXPAND | wx.ALL, 5)

        self.task_sizer.Layout()
        self.scroll.Layout()
        # refresh sidebar percentage but keep current selection
        self.load_todo_files(preserve=self.current_name)

    def add_item(self, event):
        if not self.current_file or not self.current_sheet:
            return
        dlg = ItemDialog(self, title="Tambah Task")
        if dlg.ShowModal() == wx.ID_OK:
            task, note = dlg.get_values()
            if not task:
                dlg.Destroy(); return
            wb = openpyxl.load_workbook(self.current_file)
            ws = wb[self.current_sheet]
            ws.append([task, 0, note])
            wb.save(self.current_file)
            self.show_tasks()
        dlg.Destroy()

    def edit_item(self, task_name):
        if not self.current_file or not self.current_sheet:
            return
        wb = openpyxl.load_workbook(self.current_file)
        ws = wb[self.current_sheet]
        row_idx = None
        curr_note = ""
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            t, s, n = row[:3]
            if str(t) == task_name:
                row_idx = i
                curr_note = n or ""
                break
        if not row_idx:
            return
        dlg = ItemDialog(self, title="Edit Task", task=task_name, note=curr_note)
        if dlg.ShowModal() == wx.ID_OK:
            new_task, new_note = dlg.get_values()
            if not new_task:
                dlg.Destroy(); return
            ws.cell(row=row_idx, column=1).value = new_task
            ws.cell(row=row_idx, column=3).value = new_note
            wb.save(self.current_file)
            self.show_tasks()
        dlg.Destroy()

    def delete_item(self, task_name):
        if not self.current_file or not self.current_sheet:
            return
        if wx.MessageBox(f"Hapus task '{task_name}'?", "Konfirmasi",
                         wx.YES_NO | wx.ICON_WARNING) != wx.YES:
            return
        wb = openpyxl.load_workbook(self.current_file)
        ws = wb[self.current_sheet]
        del_row = None
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            t, _, _ = row[:3]
            if str(t) == task_name:
                del_row = i
                break
        if del_row:
            ws.delete_rows(del_row, 1)
            wb.save(self.current_file)
            self.show_tasks()

    def toggle_task(self, event, task_name):
        if not self.current_file or not self.current_sheet:
            return
        new_val = 1 if event.IsChecked() else 0
        wb = openpyxl.load_workbook(self.current_file)
        ws = wb[self.current_sheet]
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == task_name:
                row[1].value = new_val
                break
        wb.save(self.current_file)
        self.show_tasks()


if __name__ == "__main__":
    app = wx.App()
    TodoApp()
    app.MainLoop()
