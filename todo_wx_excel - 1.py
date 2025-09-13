import wx
import os
import openpyxl
from openpyxl import Workbook

DATA_DIR = os.path.join("data", "todolist")

HEADER = ["Task", "Status", "Note"]  # Status: 0 = aktif, 1 = selesai

YELLOW = wx.Colour(255, 255, 102)
GREY_DONE = wx.Colour(136, 136, 136)
GREY_NOTE = wx.Colour(180, 180, 180)
BG_DARK = wx.Colour(30, 30, 30)
BG_PANEL = wx.Colour(40, 40, 40)
FG_TEXT = wx.Colour(230, 230, 230)


def ensure_data_dir():
    os.makedirs(DATA_DIR, exist_ok=True)


def create_new_todo_excel(path, first_sheet_name="Default"):
    wb = Workbook()
    ws = wb.active
    ws.title = first_sheet_name
    ws.append(HEADER)
    wb.save(path)


class ItemDialog(wx.Dialog):
    """Dialog tambah/edit item (task + note) dalam 1 kali submit."""
    def __init__(self, parent, title="Item", task="", note=""):
        super().__init__(parent, title=title, size=(420, 220))
        pnl = wx.Panel(self)
        pnl.SetBackgroundColour(BG_DARK)

        v = wx.BoxSizer(wx.VERTICAL)

        t1 = wx.StaticText(pnl, label="Nama Task")
        t1.SetForegroundColour(FG_TEXT)
        self.txt_task = wx.TextCtrl(pnl, value=task)
        self.txt_task.SetBackgroundColour(BG_PANEL)
        self.txt_task.SetForegroundColour(FG_TEXT)

        t2 = wx.StaticText(pnl, label="Keterangan (opsional)")
        t2.SetForegroundColour(FG_TEXT)
        self.txt_note = wx.TextCtrl(pnl, value=note, style=wx.TE_MULTILINE)
        self.txt_note.SetBackgroundColour(BG_PANEL)
        self.txt_note.SetForegroundColour(FG_TEXT)

        v.Add(t1, 0, wx.ALL, 6)
        v.Add(self.txt_task, 0, wx.EXPAND | wx.ALL, 6)
        v.Add(t2, 0, wx.ALL, 6)
        v.Add(self.txt_note, 1, wx.EXPAND | wx.ALL, 6)

        btns = self.CreateSeparatedButtonSizer(wx.OK | wx.CANCEL)
        v.Add(btns, 0, wx.EXPAND | wx.ALL, 6)

        pnl.SetSizer(v)
        self.Centre()

    def get_values(self):
        return self.txt_task.GetValue().strip(), self.txt_note.GetValue().strip()


class SectionDialog(wx.Dialog):
    """Dialog tambah/rename section (sheet)."""
    def __init__(self, parent, title="Section", name=""):
        super().__init__(parent, title=title, size=(360, 150))
        pnl = wx.Panel(self)
        pnl.SetBackgroundColour(BG_DARK)
        v = wx.BoxSizer(wx.VERTICAL)

        t1 = wx.StaticText(pnl, label="Nama Section")
        t1.SetForegroundColour(FG_TEXT)
        self.txt = wx.TextCtrl(pnl, value=name)
        self.txt.SetBackgroundColour(BG_PANEL)
        self.txt.SetForegroundColour(FG_TEXT)

        v.Add(t1, 0, wx.ALL, 8)
        v.Add(self.txt, 0, wx.EXPAND | wx.ALL, 8)

        btns = self.CreateSeparatedButtonSizer(wx.OK | wx.CANCEL)
        v.Add(btns, 0, wx.EXPAND | wx.ALL, 6)

        pnl.SetSizer(v)
        self.Centre()

    def get_value(self):
        return self.txt.GetValue().strip()


class RenameTodoDialog(wx.Dialog):
    """Dialog rename file Todo (tanpa ekstensi)."""
    def __init__(self, parent, old_name=""):
        super().__init__(parent, title="Rename Todo", size=(360, 150))
        pnl = wx.Panel(self)
        pnl.SetBackgroundColour(BG_DARK)
        v = wx.BoxSizer(wx.VERTICAL)

        t1 = wx.StaticText(pnl, label="Nama Todo Baru")
        t1.SetForegroundColour(FG_TEXT)
        self.txt = wx.TextCtrl(pnl, value=old_name)
        self.txt.SetBackgroundColour(BG_PANEL)
        self.txt.SetForegroundColour(FG_TEXT)

        v.Add(t1, 0, wx.ALL, 8)
        v.Add(self.txt, 0, wx.EXPAND | wx.ALL, 8)

        btns = self.CreateSeparatedButtonSizer(wx.OK | wx.CANCEL)
        v.Add(btns, 0, wx.EXPAND | wx.ALL, 6)

        pnl.SetSizer(v)
        self.Centre()

    def get_value(self):
        return self.txt.GetValue().strip()


class TodoApp(wx.Frame):
    def __init__(self):
        super().__init__(None, title="Excel Todo Manager", size=(1100, 680))
        ensure_data_dir()
        self.SetBackgroundColour(BG_DARK)

        self.file_map = {}       # display_name -> full path
        self.current_file = None # full path
        self.current_name = ""   # display name (without .xlsx)
        self.current_sheet = None

        # ===== Root layout
        root = wx.BoxSizer(wx.HORIZONTAL)

        # ===== Sidebar
        side_panel = wx.Panel(self)
        side_panel.SetBackgroundColour(BG_PANEL)
        side_sizer = wx.BoxSizer(wx.VERTICAL)

        lbl = wx.StaticText(side_panel, label="Daftar Todo")
        lbl.SetForegroundColour(FG_TEXT)
        font = lbl.GetFont()
        font.MakeBold()
        lbl.SetFont(font)

        self.todo_list = wx.ListBox(side_panel, size=(280, -1))
        self.todo_list.Bind(wx.EVT_LISTBOX, self.on_select_todo)
        self.todo_list.SetBackgroundColour(BG_DARK)
        self.todo_list.SetForegroundColour(FG_TEXT)

        btn_new = wx.Button(side_panel, label="Buat Todo Baru")
        btn_ren = wx.Button(side_panel, label="Rename Todo")
        btn_del = wx.Button(side_panel, label="Hapus Todo")
        btn_ref = wx.Button(side_panel, label="Refresh")

        for b in (btn_new, btn_ren, btn_del, btn_ref):
            b.SetBackgroundColour(wx.Colour(0, 122, 204))
            b.SetForegroundColour(wx.WHITE)

        btn_new.Bind(wx.EVT_BUTTON, self.create_todo)
        btn_ren.Bind(wx.EVT_BUTTON, self.rename_todo)
        btn_del.Bind(wx.EVT_BUTTON, self.delete_todo)
        btn_ref.Bind(wx.EVT_BUTTON, lambda e: self.load_todo_files())

        side_sizer.Add(lbl, 0, wx.ALL, 8)
        side_sizer.Add(self.todo_list, 1, wx.EXPAND | wx.ALL, 8)
        side_sizer.Add(btn_new, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 6)
        side_sizer.Add(btn_ren, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 6)
        side_sizer.Add(btn_del, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 6)
        side_sizer.Add(btn_ref, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 6)
        side_panel.SetSizer(side_sizer)

        root.Add(side_panel, 0, wx.EXPAND)

        # ===== Main area
        main_panel = wx.Panel(self)
        main_panel.SetBackgroundColour(BG_DARK)
        main_sizer = wx.BoxSizer(wx.VERTICAL)

        # Top bar: current todo name
        self.lbl_title = wx.StaticText(main_panel, label="(Belum memilih Todo)")
        self.lbl_title.SetForegroundColour(FG_TEXT)
        ft = self.lbl_title.GetFont()
        ft.SetPointSize(ft.GetPointSize() + 2)
        ft.MakeBold()
        self.lbl_title.SetFont(ft)
        main_sizer.Add(self.lbl_title, 0, wx.ALL, 8)

        # Sheet controls
        sheet_bar = wx.BoxSizer(wx.HORIZONTAL)
        self.sheet_list = wx.Choice(main_panel, choices=[])
        self.sheet_list.Bind(wx.EVT_CHOICE, self.on_select_sheet)
        self.sheet_list.SetBackgroundColour(BG_PANEL)
        self.sheet_list.SetForegroundColour(FG_TEXT)

        btn_add_sec = wx.Button(main_panel, label="Tambah Section")
        btn_ren_sec = wx.Button(main_panel, label="Rename Section")
        btn_del_sec = wx.Button(main_panel, label="Hapus Section")
        for b in (btn_add_sec, btn_ren_sec, btn_del_sec):
            b.SetBackgroundColour(wx.Colour(85, 85, 85))
            b.SetForegroundColour(wx.WHITE)

        btn_add_sec.Bind(wx.EVT_BUTTON, self.add_section)
        btn_ren_sec.Bind(wx.EVT_BUTTON, self.rename_section)
        btn_del_sec.Bind(wx.EVT_BUTTON, self.delete_section)

        sheet_bar.Add(wx.StaticText(main_panel, label="Section:"), 0, wx.ALIGN_CENTER_VERTICAL | wx.RIGHT, 6)
        sheet_bar.Add(self.sheet_list, 0, wx.RIGHT, 12)
        sheet_bar.Add(btn_add_sec, 0, wx.RIGHT, 6)
        sheet_bar.Add(btn_ren_sec, 0, wx.RIGHT, 6)
        sheet_bar.Add(btn_del_sec, 0)

        main_sizer.Add(sheet_bar, 0, wx.ALL, 8)

        # Task toolbar
        task_bar = wx.BoxSizer(wx.HORIZONTAL)
        btn_add_item = wx.Button(main_panel, label="Tambah Task")
        btn_add_item.SetBackgroundColour(wx.Colour(0, 122, 204))
        btn_add_item.SetForegroundColour(wx.WHITE)
        btn_add_item.Bind(wx.EVT_BUTTON, self.add_item)

        task_bar.Add(btn_add_item, 0, wx.RIGHT, 6)
        main_sizer.Add(task_bar, 0, wx.LEFT | wx.RIGHT | wx.BOTTOM, 8)

        # Scrolled task area
        self.scroll = wx.ScrolledWindow(main_panel, style=wx.VSCROLL)
        self.scroll.SetScrollRate(0, 14)
        self.scroll.SetBackgroundColour(BG_DARK)
        self.task_area = wx.Panel(self.scroll)
        self.task_area.SetBackgroundColour(BG_DARK)
        self.task_sizer = wx.BoxSizer(wx.VERTICAL)
        self.task_area.SetSizer(self.task_sizer)

        sv = wx.BoxSizer(wx.VERTICAL)
        sv.Add(self.task_area, 1, wx.EXPAND | wx.ALL, 6)
        self.scroll.SetSizer(sv)

        main_sizer.Add(self.scroll, 1, wx.EXPAND | wx.ALL, 8)

        main_panel.SetSizer(main_sizer)
        root.Add(main_panel, 1, wx.EXPAND)

        self.SetSizer(root)
        self.Centre()
        self.load_todo_files()
        self.Show()

        # Re-wrap texts on resize
        self.Bind(wx.EVT_SIZE, lambda e: self.wrap_all_texts() or e.Skip())

    # ===== Sidebar & Todo files
    def load_todo_files(self):
        """Load semua file Excel di DATA_DIR ke sidebar, tampil tanpa .xlsx."""
        self.file_map.clear()
        display_names = []
        for f in sorted(os.listdir(DATA_DIR)):
            if f.lower().endswith(".xlsx"):
                name = os.path.splitext(f)[0]
                self.file_map[name] = os.path.join(DATA_DIR, f)
                display_names.append(name)
        self.todo_list.Set(display_names)

    def create_todo(self, event):
        dlg = RenameTodoDialog(self, old_name="")
        if dlg.ShowModal() == wx.ID_OK:
            name = dlg.get_value()
            if not name:
                wx.MessageBox("Nama tidak boleh kosong.", "Error", wx.OK | wx.ICON_ERROR)
                return
            path = os.path.join(DATA_DIR, f"{name}.xlsx")
            if os.path.exists(path):
                wx.MessageBox("Nama sudah digunakan.", "Error", wx.OK | wx.ICON_ERROR)
                return
            create_new_todo_excel(path, first_sheet_name="Default")
            self.load_todo_files()
        dlg.Destroy()

    def rename_todo(self, event):
        sel = self.todo_list.GetSelection()
        if sel == wx.NOT_FOUND:
            wx.MessageBox("Pilih todo dahulu.", "Info", wx.OK | wx.ICON_INFORMATION)
            return
        old_display = self.todo_list.GetString(sel)
        old_path = self.file_map.get(old_display)
        dlg = RenameTodoDialog(self, old_name=old_display)
        if dlg.ShowModal() == wx.ID_OK:
            new_name = dlg.get_value()
            if not new_name:
                wx.MessageBox("Nama tidak boleh kosong.", "Error", wx.OK | wx.ICON_ERROR)
                return
            new_path = os.path.join(DATA_DIR, f"{new_name}.xlsx")
            if os.path.exists(new_path):
                wx.MessageBox("Nama sudah digunakan.", "Error", wx.OK | wx.ICON_ERROR)
                return
            os.rename(old_path, new_path)
            self.load_todo_files()
            # perbarui pilihan jika yang di-rename sedang aktif
            if self.current_file == old_path:
                self.current_file = new_path
                self.current_name = new_name
                self.lbl_title.SetLabel(f"# {self.current_name}")
        dlg.Destroy()

    def delete_todo(self, event):
        sel = self.todo_list.GetSelection()
        if sel == wx.NOT_FOUND:
            wx.MessageBox("Pilih todo dahulu.", "Info", wx.OK | wx.ICON_INFORMATION)
            return
        display = self.todo_list.GetString(sel)
        path = self.file_map.get(display)
        if wx.MessageBox(f"Hapus todo '{display}'?", "Konfirmasi", wx.YES_NO | wx.ICON_WARNING) == wx.YES:
            try:
                os.remove(path)
            except Exception as e:
                wx.MessageBox(str(e), "Error", wx.OK | wx.ICON_ERROR)
                return
            self.current_file = None
            self.current_name = ""
            self.current_sheet = None
            self.lbl_title.SetLabel("(Belum memilih Todo)")
            self.sheet_list.Set([])
            self.clear_tasks()
            self.load_todo_files()

    def on_select_todo(self, event):
        display = self.todo_list.GetStringSelection()
        path = self.file_map.get(display)
        if not path:
            return
        self.current_file = path
        self.current_name = display
        self.lbl_title.SetLabel(f"# {self.current_name}")
        self.load_sheets()

    # ===== Sheets (Sections)
    def load_sheets(self):
        try:
            wb = openpyxl.load_workbook(self.current_file)
        except Exception as e:
            wx.MessageBox(str(e), "Error", wx.OK | wx.ICON_ERROR)
            return
        names = wb.sheetnames
        self.sheet_list.Set(names)
        if names:
            # pilih sheet pertama jika belum dipilih
            idx = 0
            if self.current_sheet in names:
                idx = names.index(self.current_sheet)
            self.sheet_list.SetSelection(idx)
            self.current_sheet = names[idx]
            self.show_tasks()
        else:
            self.current_sheet = None
            self.clear_tasks()

    def add_section(self, event):
        if not self.current_file:
            wx.MessageBox("Pilih todo dahulu.", "Info", wx.OK | wx.ICON_INFORMATION)
            return
        dlg = SectionDialog(self, title="Tambah Section")
        if dlg.ShowModal() == wx.ID_OK:
            name = dlg.get_value()
            if not name:
                wx.MessageBox("Nama section tidak boleh kosong.", "Error", wx.OK | wx.ICON_ERROR)
                return
            wb = openpyxl.load_workbook(self.current_file)
            if name in wb.sheetnames:
                wx.MessageBox("Nama section sudah ada.", "Error", wx.OK | wx.ICON_ERROR)
                return
            ws = wb.create_sheet(name)
            ws.append(HEADER)
            wb.save(self.current_file)
            self.current_sheet = name
            self.load_sheets()
        dlg.Destroy()

    def rename_section(self, event):
        if not self.current_file or not self.current_sheet:
            wx.MessageBox("Pilih todo & section dahulu.", "Info", wx.OK | wx.ICON_INFORMATION)
            return
        dlg = SectionDialog(self, title="Rename Section", name=self.current_sheet)
        if dlg.ShowModal() == wx.ID_OK:
            new_name = dlg.get_value()
            if not new_name:
                wx.MessageBox("Nama section tidak boleh kosong.", "Error", wx.OK | wx.ICON_ERROR)
                return
            wb = openpyxl.load_workbook(self.current_file)
            if new_name in wb.sheetnames:
                wx.MessageBox("Nama section sudah ada.", "Error", wx.OK | wx.ICON_ERROR)
                return
            ws = wb[self.current_sheet]
            ws.title = new_name
            wb.save(self.current_file)
            self.current_sheet = new_name
            self.load_sheets()
        dlg.Destroy()

    def on_select_sheet(self, event):
        """Handler saat user memilih section dari dropdown."""
        self.current_sheet = self.sheet_list.GetStringSelection()
        self.show_tasks()

    def delete_section(self, event):
        if not self.current_file or not self.current_sheet:
            wx.MessageBox("Pilih todo & section dahulu.", "Info", wx.OK | wx.ICON_INFORMATION)
            return
        wb = openpyxl.load_workbook(self.current_file)
        if len(wb.sheetnames) <= 1:
            wx.MessageBox("Minimal harus ada 1 section.", "Info", wx.OK | wx.ICON_INFORMATION)
            return
        if wx.MessageBox(f"Hapus section '{self.current_sheet}'?", "Konfirmasi",
                         wx.YES_NO | wx.ICON_WARNING) == wx.YES:
            ws = wb[self.current_sheet]
            wb.remove(ws)
            wb.save(self.current_file)
            self.current_sheet = None
            self.load_sheets()

    # ===== Tasks (Items)
    def clear_tasks(self):
        for c in self.task_area.GetChildren():
            c.Destroy()
        self.task_sizer.Layout()
        self.scroll.Layout()

    def show_tasks(self):
        self.clear_tasks()
        if not self.current_file or not self.current_sheet:
            return

        wb = openpyxl.load_workbook(self.current_file)
        ws = wb[self.current_sheet]

        # pastikan header ada
        if ws.max_row == 0:
            ws.append(HEADER)
            wb.save(self.current_file)

        # generate rows UI
        for row in ws.iter_rows(min_row=2, values_only=False):
            task_cell, status_cell, note_cell = row[:3]
            task = (task_cell.value or "").strip() if task_cell.value else ""
            status = int(status_cell.value) if status_cell.value in (0, 1) else int(bool(status_cell.value or 0))
            note = (note_cell.value or "").strip() if note_cell and note_cell.value else ""

            row_panel = wx.Panel(self.task_area)
            row_panel.SetBackgroundColour(BG_PANEL)
            vs = wx.BoxSizer(wx.VERTICAL)

            # Bar atas: checkbox + tombol edit/hapus
            top_bar = wx.BoxSizer(wx.HORIZONTAL)
            cb = wx.CheckBox(row_panel, label=task)
            cb.SetValue(bool(status))
            cb.SetForegroundColour(GREY_DONE if status else YELLOW)
            cb.Bind(wx.EVT_CHECKBOX, lambda e, t=task: self.toggle_task(e, t))

            btn_edit = wx.Button(row_panel, label="Edit", size=(70, 26))
            btn_del = wx.Button(row_panel, label="Hapus", size=(70, 26))
            btn_edit.SetBackgroundColour(wx.Colour(85, 85, 85))
            btn_edit.SetForegroundColour(wx.WHITE)
            btn_del.SetBackgroundColour(wx.Colour(170, 0, 0))
            btn_del.SetForegroundColour(wx.WHITE)
            btn_edit.Bind(wx.EVT_BUTTON, lambda e, t=task: self.edit_item(t))
            btn_del.Bind(wx.EVT_BUTTON, lambda e, t=task: self.delete_item(t))

            top_bar.Add(cb, 1, wx.ALIGN_CENTER_VERTICAL | wx.ALL, 6)
            top_bar.Add(btn_edit, 0, wx.ALL, 6)
            top_bar.Add(btn_del, 0, wx.RIGHT | wx.TOP | wx.BOTTOM, 6)

            vs.Add(top_bar, 0, wx.EXPAND)

            # Note di bawah (wrap)
            if note:
                txt = wx.StaticText(row_panel, label=note)
                txt.SetForegroundColour(GREY_NOTE)
                vs.Add(txt, 0, wx.LEFT | wx.RIGHT | wx.BOTTOM, 10)

                # simpan untuk dibungkus saat resize
                row_panel._wrap_targets = [cb, txt]
            else:
                row_panel._wrap_targets = [cb]

            row_panel.SetSizer(vs)
            self.task_sizer.Add(row_panel, 0, wx.EXPAND | wx.ALL, 6)

        self.task_sizer.Layout()
        self.scroll.Layout()
        self.wrap_all_texts()

    def add_item(self, event):
        if not self.current_file or not self.current_sheet:
            wx.MessageBox("Pilih todo & section dahulu.", "Info", wx.OK | wx.ICON_INFORMATION)
            return
        dlg = ItemDialog(self, title="Tambah Task")
        if dlg.ShowModal() == wx.ID_OK:
            task, note = dlg.get_values()
            if not task:
                wx.MessageBox("Nama task tidak boleh kosong.", "Error", wx.OK | wx.ICON_ERROR)
                dlg.Destroy()
                return
            wb = openpyxl.load_workbook(self.current_file)
            ws = wb[self.current_sheet]
            ws.append([task, 0, note])
            wb.save(self.current_file)
            self.show_tasks()
        dlg.Destroy()

    def edit_item(self, task_name):
        if not self.current_file or not self.current_sheet:
            return
        wb = openpyxl.load_workbook(self.current_file)
        ws = wb[self.current_sheet]

        row_idx = None
        curr_note = ""
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            t, s, n = row[:3]
            if (t or "") == task_name:
                row_idx = i
                curr_note = n or ""
                break
        if row_idx is None:
            return

        dlg = ItemDialog(self, title="Edit Task", task=task_name, note=curr_note)
        if dlg.ShowModal() == wx.ID_OK:
            new_task, new_note = dlg.get_values()
            if not new_task:
                wx.MessageBox("Nama task tidak boleh kosong.", "Error", wx.OK | wx.ICON_ERROR)
                dlg.Destroy()
                return
            ws.cell(row=row_idx, column=1).value = new_task
            ws.cell(row=row_idx, column=3).value = new_note
            wb.save(self.current_file)
            self.show_tasks()
        dlg.Destroy()

    def delete_item(self, task_name):
        if not self.current_file or not self.current_sheet:
            return
        if wx.MessageBox(f"Hapus task '{task_name}'?", "Konfirmasi", wx.YES_NO | wx.ICON_WARNING) != wx.YES:
            return
        wb = openpyxl.load_workbook(self.current_file)
        ws = wb[self.current_sheet]

        del_row = None
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            t, _, _ = row[:3]
            if (t or "") == task_name:
                del_row = i
                break
        if del_row:
            ws.delete_rows(del_row, 1)
            wb.save(self.current_file)
            self.show_tasks()

    def toggle_task(self, event, task_name):
        if not self.current_file or not self.current_sheet:
            return
        new_val = 1 if event.IsChecked() else 0
        wb = openpyxl.load_workbook(self.current_file)
        ws = wb[self.current_sheet]
        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if (row[0].value or "") == task_name:
                row[1].value = new_val
                break
        wb.save(self.current_file)
        self.show_tasks()

    # ===== Wrap texts when resizing
    def wrap_all_texts(self):
        """Wrap label teks sesuai lebar area task."""
        width = self.scroll.GetClientSize().GetWidth() - 40
        if width < 100:
            width = 100
        for row_panel in self.task_area.GetChildren():
            targets = getattr(row_panel, "_wrap_targets", [])
            for w in targets:
                if isinstance(w, wx.CheckBox):
                    # untuk CheckBox, kita pakai label terpisah? tidak, wx tidak punya Wrap di checkbox
                    # solusi: biarkan satu baris; untuk teks panjang sebaiknya gunakan StaticText.
                    pass
                elif isinstance(w, wx.StaticText):
                    w.Wrap(width)
        self.task_sizer.Layout()
        self.scroll.Layout()


if __name__ == "__main__":
    app = wx.App()
    TodoApp()
    app.MainLoop()
