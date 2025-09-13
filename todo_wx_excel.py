import wx
import wx.adv
import os
import openpyxl
import datetime
from openpyxl import Workbook

# ===== Config & Colors =====
DATA_DIR = os.path.join("data", "todolist")
HEADER = ["Task", "Status", "Note", "Tanggal"]  # Status: 0 = aktif, 1 = selesai

YELLOW = wx.Colour(255, 255, 102)
GREY_DONE = wx.Colour(136, 136, 136)
GREY_NOTE = wx.Colour(180, 180, 180)
BG_DARK = wx.Colour(30, 30, 30)
BG_PANEL = wx.Colour(40, 40, 40)
FG_TEXT = wx.Colour(230, 230, 230)


# ===== Helpers =====
def ensure_data_dir():
    os.makedirs(DATA_DIR, exist_ok=True)


def create_new_todo_excel(path, first_sheet_name="Default"):
    wb = Workbook()
    ws = wb.active
    ws.title = first_sheet_name
    ws.append(HEADER)
    wb.save(path)


def safe_unpack(row):
    values = list(row)
    if len(values) < 4:
        values += [""] * (4 - len(values))
    return values[:4]


def calc_todo_progress(path) -> int:
    try:
        wb = openpyxl.load_workbook(path)
    except Exception:
        return 0
    total, done = 0, 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=2, values_only=True):
            task, status, note, tanggal = safe_unpack(row)
            if task and str(task).strip() != "":
                total += 1
                if status in (1, "1", True):
                    done += 1
    if total == 0:
        return 0
    return int((done / total) * 100)


def parse_display_to_name(display_text: str) -> str:
    if " (" in display_text and display_text.endswith(")"):
        return display_text.rsplit(" (", 1)[0]
    return display_text


# ===== Dialogs =====
class ItemDialog(wx.Dialog):
    def __init__(self, parent, title="Item", task="", note="", tanggal=None):
        super().__init__(parent, title=title, size=(430, 350))
        pnl = wx.Panel(self)
        pnl.SetBackgroundColour(BG_DARK)

        v = wx.BoxSizer(wx.VERTICAL)

        t1 = wx.StaticText(pnl, label="Nama Task")
        t1.SetForegroundColour(FG_TEXT)
        self.txt_task = wx.TextCtrl(pnl, value=task)

        t2 = wx.StaticText(pnl, label="Keterangan")
        t2.SetForegroundColour(FG_TEXT)
        self.txt_note = wx.TextCtrl(pnl, value=note, style=wx.TE_MULTILINE)

        t3 = wx.StaticText(pnl, label="Tanggal")
        t3.SetForegroundColour(FG_TEXT)
        self.date_picker = wx.adv.DatePickerCtrl(
            pnl, style=wx.adv.DP_DROPDOWN | wx.adv.DP_SHOWCENTURY
        )

        # default
        if tanggal:
            try:
                y, m, d = map(int, str(tanggal).split("-"))
                self.date_picker.SetValue(wx.DateTime.FromDMY(d, m - 1, y))
            except:
                pass
        else:
            today = datetime.date.today()
            self.date_picker.SetValue(
                wx.DateTime.FromDMY(today.day, today.month - 1, today.year)
            )

        btns = wx.StdDialogButtonSizer()
        ok_btn = wx.Button(pnl, wx.ID_OK)
        cancel_btn = wx.Button(pnl, wx.ID_CANCEL)
        btns.AddButton(ok_btn)
        btns.AddButton(cancel_btn)
        btns.Realize()

        v.Add(t1, 0, wx.ALL, 8)
        v.Add(self.txt_task, 0, wx.EXPAND | wx.LEFT | wx.RIGHT | wx.BOTTOM, 8)
        v.Add(t2, 0, wx.LEFT | wx.RIGHT | wx.TOP, 8)
        v.Add(self.txt_note, 1, wx.EXPAND | wx.ALL, 8)
        v.Add(t3, 0, wx.LEFT | wx.RIGHT | wx.TOP, 8)
        v.Add(self.date_picker, 0, wx.EXPAND | wx.ALL, 8)
        v.Add(btns, 0, wx.EXPAND | wx.ALL, 8)

        pnl.SetSizer(v)
        self.Centre()

    def get_values(self):
        task = self.txt_task.GetValue().strip()
        note = self.txt_note.GetValue().strip()
        d = self.date_picker.GetValue()
        tanggal = f"{d.GetYear()}-{d.GetMonth()+1:02d}-{d.GetDay():02d}"
        return task, note, tanggal


class SectionDialog(wx.Dialog):
    def __init__(self, parent, title="Section", name=""):
        super().__init__(parent, title=title, size=(360, 160))
        pnl = wx.Panel(self); pnl.SetBackgroundColour(BG_DARK)
        v = wx.BoxSizer(wx.VERTICAL)
        t1 = wx.StaticText(pnl, label="Nama Section"); t1.SetForegroundColour(FG_TEXT)
        self.txt = wx.TextCtrl(pnl, value=name)
        btns = wx.StdDialogButtonSizer()
        ok_btn = wx.Button(pnl, wx.ID_OK); cancel_btn = wx.Button(pnl, wx.ID_CANCEL)
        btns.AddButton(ok_btn); btns.AddButton(cancel_btn); btns.Realize()
        v.Add(t1, 0, wx.ALL, 10); v.Add(self.txt, 0, wx.EXPAND | wx.ALL, 10); v.Add(btns, 0, wx.EXPAND | wx.ALL, 10)
        pnl.SetSizer(v); self.Centre()
    def get_value(self): return self.txt.GetValue().strip()


class RenameTodoDialog(wx.Dialog):
    def __init__(self, parent, old_name=""):
        super().__init__(parent, title="Nama Todo", size=(360, 160))
        pnl = wx.Panel(self); pnl.SetBackgroundColour(BG_DARK)
        v = wx.BoxSizer(wx.VERTICAL)
        t1 = wx.StaticText(pnl, label="Nama Todo"); t1.SetForegroundColour(FG_TEXT)
        self.txt = wx.TextCtrl(pnl, value=old_name)
        btns = wx.StdDialogButtonSizer()
        ok_btn = wx.Button(pnl, wx.ID_OK); cancel_btn = wx.Button(pnl, wx.ID_CANCEL)
        btns.AddButton(ok_btn); btns.AddButton(cancel_btn); btns.Realize()
        v.Add(t1, 0, wx.ALL, 10); v.Add(self.txt, 0, wx.EXPAND | wx.ALL, 10); v.Add(btns, 0, wx.EXPAND | wx.ALL, 10)
        pnl.SetSizer(v); self.Centre()
    def get_value(self): return self.txt.GetValue().strip()


# ===== Main App =====
class TodoApp(wx.Frame):
    def __init__(self):
        super().__init__(None, title="Excel Todo Manager", size=(1150, 720))
        ensure_data_dir(); self.SetBackgroundColour(BG_DARK)
        self.file_map = {}; self.current_file=None; self.current_name=""; self.current_sheet=None

        root = wx.BoxSizer(wx.HORIZONTAL)

        # Sidebar
        side_panel = wx.Panel(self); side_panel.SetBackgroundColour(BG_PANEL)
        side_sizer = wx.BoxSizer(wx.VERTICAL)
        lbl = wx.StaticText(side_panel, label="Daftar Todo"); lbl.SetForegroundColour(FG_TEXT); f=lbl.GetFont(); f.MakeBold(); lbl.SetFont(f)
        self.todo_list = wx.ListBox(side_panel, size=(300,-1)); self.todo_list.Bind(wx.EVT_LISTBOX, self.on_select_todo)
        self.todo_list.SetBackgroundColour(BG_DARK); self.todo_list.SetForegroundColour(FG_TEXT)
        btn_new=wx.Button(side_panel,label="Buat Todo Baru"); btn_ren=wx.Button(side_panel,label="Rename Todo"); btn_del=wx.Button(side_panel,label="Hapus Todo"); btn_ref=wx.Button(side_panel,label="Refresh")
        for b in (btn_new,btn_ren,btn_del,btn_ref): b.SetBackgroundColour(wx.Colour(0,122,204)); b.SetForegroundColour(wx.WHITE)
        btn_new.Bind(wx.EVT_BUTTON,self.create_todo); btn_ren.Bind(wx.EVT_BUTTON,self.rename_todo); btn_del.Bind(wx.EVT_BUTTON,self.delete_todo); btn_ref.Bind(wx.EVT_BUTTON,lambda e:self.load_todo_files(preserve=self.current_name))
        side_sizer.Add(lbl,0,wx.ALL,8); side_sizer.Add(self.todo_list,1,wx.EXPAND|wx.ALL,8)
        for b in (btn_new,btn_ren,btn_del,btn_ref): side_sizer.Add(b,0,wx.EXPAND|wx.ALL,6)
        side_panel.SetSizer(side_sizer); root.Add(side_panel,0,wx.EXPAND)

        # Main area
        main_panel=wx.Panel(self); main_panel.SetBackgroundColour(BG_DARK); main_sizer=wx.BoxSizer(wx.VERTICAL)
        self.lbl_title=wx.StaticText(main_panel,label="(Belum memilih Todo)"); self.lbl_title.SetForegroundColour(FG_TEXT)
        ft=self.lbl_title.GetFont(); ft.MakeBold(); ft.SetPointSize(ft.GetPointSize()+2); self.lbl_title.SetFont(ft); main_sizer.Add(self.lbl_title,0,wx.ALL,10)

        bar=wx.BoxSizer(wx.HORIZONTAL)
        bar.Add(wx.StaticText(main_panel,label="Section:"),0,wx.ALIGN_CENTER_VERTICAL|wx.RIGHT,6)
        self.sheet_list=wx.Choice(main_panel,choices=[]); self.sheet_list.Bind(wx.EVT_CHOICE,self.on_select_sheet); self.sheet_list.SetBackgroundColour(BG_PANEL); self.sheet_list.SetForegroundColour(FG_TEXT); bar.Add(self.sheet_list,0,wx.RIGHT,12)
        btn_add_sec=wx.Button(main_panel,label="Tambah Section"); btn_ren_sec=wx.Button(main_panel,label="Rename Section"); btn_del_sec=wx.Button(main_panel,label="Hapus Section")
        for b in (btn_add_sec,btn_ren_sec,btn_del_sec): b.SetBackgroundColour(wx.Colour(85,85,85)); b.SetForegroundColour(wx.WHITE)
        btn_add_sec.Bind(wx.EVT_BUTTON,self.add_section); btn_ren_sec.Bind(wx.EVT_BUTTON,self.rename_section); btn_del_sec.Bind(wx.EVT_BUTTON,self.delete_section)
        for b in (btn_add_sec,btn_ren_sec,btn_del_sec): bar.Add(b,0,wx.RIGHT,6)
        bar.Add(wx.StaticText(main_panel,label="Filter Tanggal:"),0,wx.ALIGN_CENTER_VERTICAL|wx.RIGHT,6)
        self.date_filter=wx.Choice(main_panel,choices=["Semua Tanggal"]); self.date_filter.Bind(wx.EVT_CHOICE,lambda e:self.show_tasks()); bar.Add(self.date_filter,0)
        main_sizer.Add(bar,0,wx.ALL,10)

        btn_add_item=wx.Button(main_panel,label="Tambah Task"); btn_add_item.SetBackgroundColour(wx.Colour(0,122,204)); btn_add_item.SetForegroundColour(wx.WHITE); btn_add_item.Bind(wx.EVT_BUTTON,self.add_item)
        main_sizer.Add(btn_add_item,0,wx.LEFT|wx.RIGHT|wx.BOTTOM,10)

        self.scroll=wx.ScrolledWindow(main_panel,style=wx.VSCROLL); self.scroll.SetScrollRate(0,14); self.scroll.SetBackgroundColour(BG_DARK)
        self.task_area=wx.Panel(self.scroll); self.task_area.SetBackgroundColour(BG_DARK); self.task_sizer=wx.BoxSizer(wx.VERTICAL); self.task_area.SetSizer(self.task_sizer)
        sv=wx.BoxSizer(wx.VERTICAL); sv.Add(self.task_area,1,wx.EXPAND|wx.ALL,6); self.scroll.SetSizer(sv); main_sizer.Add(self.scroll,1,wx.EXPAND|wx.ALL,10)

        main_panel.SetSizer(main_sizer); root.Add(main_panel,1,wx.EXPAND)
        self.SetSizer(root); self.Centre(); self.load_todo_files(); self.Show()

    # === CRUD Todo (file)
    def load_todo_files(self,preserve=None):
        self.file_map.clear(); displays=[]
        for f in sorted(os.listdir(DATA_DIR)):
            if f.lower().endswith(".xlsx"):
                name=os.path.splitext(f)[0]; path=os.path.join(DATA_DIR,f); self.file_map[name]=path
                pct=calc_todo_progress(path); displays.append(f"{name} ({pct}%)")
        self.todo_list.Set(displays)
        if preserve:
            for i,txt in enumerate(self.todo_list.GetStrings()):
                if parse_display_to_name(txt)==preserve: self.todo_list.SetSelection(i); break
    def get_selected_todo_name(self):
        sel=self.todo_list.GetSelection(); 
        if sel==wx.NOT_FOUND: return None
        return parse_display_to_name(self.todo_list.GetString(sel))
    def create_todo(self,e):
        dlg=RenameTodoDialog(self); 
        if dlg.ShowModal()==wx.ID_OK:
            name=dlg.get_value(); 
            if not name: return
            path=os.path.join(DATA_DIR,f"{name}.xlsx"); 
            if os.path.exists(path): wx.MessageBox("Nama sudah ada","Error"); return
            create_new_todo_excel(path); self.load_todo_files(preserve=name)
        dlg.Destroy()
    def rename_todo(self,e):
        name=self.get_selected_todo_name(); 
        if not name: return
        dlg=RenameTodoDialog(self,old_name=name); 
        if dlg.ShowModal()==wx.ID_OK:
            new_name=dlg.get_value(); 
            if not new_name: return
            old_path=self.file_map[name]; new_path=os.path.join(DATA_DIR,f"{new_name}.xlsx"); os.rename(old_path,new_path)
            if self.current_file==old_path: self.current_file=new_path; self.current_name=new_name; self.lbl_title.SetLabel(f"# {new_name}")
            self.load_todo_files(preserve=new_name)
        dlg.Destroy()
    def delete_todo(self,e):
        name=self.get_selected_todo_name(); 
        if not name: return
        path=self.file_map[name]; 
        if wx.MessageBox(f"Hapus todo '{name}'?","Konfirmasi",wx.YES_NO)==wx.YES:
            os.remove(path); 
            if self.current_name==name: self.current_file=None; self.current_sheet=None; self.lbl_title.SetLabel("(Belum memilih Todo)"); self.sheet_list.Set([]); self.clear_tasks()
            self.load_todo_files()
    def on_select_todo(self,e):
        name=self.get_selected_todo_name(); 
        if not name: return
        path=self.file_map[name]; self.current_file=path; self.current_name=name; self.lbl_title.SetLabel(f"# {name}"); self.load_sheets()

    # === CRUD Section
    def on_select_sheet(self,e): self.current_sheet=self.sheet_list.GetStringSelection(); self.show_tasks()
    def load_sheets(self):
        wb=openpyxl.load_workbook(self.current_file); names=wb.sheetnames; self.sheet_list.Set(names)
        if not names: return
        self.current_sheet=names[0]; self.sheet_list.SetSelection(0); self.show_tasks()
    def add_section(self,e):
        if not self.current_file: return
        dlg=SectionDialog(self,title="Tambah Section"); 
        if dlg.ShowModal()==wx.ID_OK:
            name=dlg.get_value(); 
            if not name: return
            wb=openpyxl.load_workbook(self.current_file); 
            if name in wb.sheetnames: wx.MessageBox("Section sudah ada","Error"); return
            ws=wb.create_sheet(name); ws.append(HEADER); wb.save(self.current_file); self.current_sheet=name; self.load_sheets()
        dlg.Destroy()
    def rename_section(self,e):
        if not self.current_file or not self.current_sheet: return
        dlg=SectionDialog(self,title="Rename Section",name=self.current_sheet); 
        if dlg.ShowModal()==wx.ID_OK:
            new_name=dlg.get_value(); 
            if not new_name: return
            wb=openpyxl.load_workbook(self.current_file); 
            if new_name in wb.sheetnames: wx.MessageBox("Section sudah ada","Error"); return
            ws=wb[self.current_sheet]; ws.title=new_name; wb.save(self.current_file); self.current_sheet=new_name; self.load_sheets()
        dlg.Destroy()
    def delete_section(self,e):
        if not self.current_file or not self.current_sheet: return
        wb=openpyxl.load_workbook(self.current_file); 
        if len(wb.sheetnames)<=1: wx.MessageBox("Minimal 1 section","Info"); return
        ws=wb[self.current_sheet]; wb.remove(ws); wb.save(self.current_file); self.current_sheet=None; self.load_sheets()

    # === CRUD Task
    def clear_tasks(self): [c.Destroy() for c in self.task_area.GetChildren()]; self.task_sizer.Layout(); self.scroll.Layout()
    def refresh_date_filter(self,ws,keep_selection=None):
        dates=set()
        for row in ws.iter_rows(min_row=2,values_only=True):
            _,_,_,tanggal=safe_unpack(row)
            if tanggal: dates.add(str(tanggal)[:7])
        items=["Semua Tanggal"]+sorted(dates); self.date_filter.Set(items)
        if keep_selection and keep_selection in items: self.date_filter.SetStringSelection(keep_selection)
        else: self.date_filter.SetSelection(0)
    def show_tasks(self):
        self.clear_tasks()
        if not self.current_file or not self.current_sheet: return
        wb=openpyxl.load_workbook(self.current_file); ws=wb[self.current_sheet]
        if ws.max_row==0: ws.append(HEADER); wb.save(self.current_file)
        selected_filter=self.date_filter.GetStringSelection()
        self.refresh_date_filter(ws,keep_selection=selected_filter)
        filter_val=self.date_filter.GetStringSelection()
        for row in ws.iter_rows(min_row=2,values_only=True):
            task,status,note,tanggal=safe_unpack(row)
            if not task: continue
            status=int(status) if status else 0; note=note or ""; tanggal=tanggal or ""
            if filter_val!="Semua Tanggal" and not str(tanggal).startswith(filter_val): continue
            row_panel=wx.Panel(self.task_area); row_panel.SetBackgroundColour(BG_PANEL); vs=wx.BoxSizer(wx.VERTICAL)
            top=wx.BoxSizer(wx.HORIZONTAL)
            cb=wx.CheckBox(row_panel,label=f"{task} [{tanggal}]"); cb.SetValue(bool(status))
            cb.SetForegroundColour(GREY_DONE if status else YELLOW); cb.Bind(wx.EVT_CHECKBOX,lambda e,t=str(task):self.toggle_task(e,t))
            btn_edit=wx.Button(row_panel,label="Edit",size=(72,26)); btn_del=wx.Button(row_panel,label="Hapus",size=(72,26))
            btn_edit.Bind(wx.EVT_BUTTON,lambda e,t=str(task):self.edit_item(t)); btn_del.Bind(wx.EVT_BUTTON,lambda e,t=str(task):self.delete_item(t))
            top.Add(cb,1,wx.ALL|wx.ALIGN_CENTER_VERTICAL,6); top.Add(btn_edit,0,wx.ALL,6); top.Add(btn_del,0,wx.ALL,6); vs.Add(top,0,wx.EXPAND)
            if note: txt=wx.StaticText(row_panel,label=str(note)); txt.SetForegroundColour(GREY_NOTE); vs.Add(txt,0,wx.LEFT|wx.RIGHT|wx.BOTTOM,10)
            row_panel.SetSizer(vs); self.task_sizer.Add(row_panel,0,wx.EXPAND|wx.ALL,5)
        self.task_sizer.Layout(); self.scroll.Layout(); self.load_todo_files(preserve=self.current_name)
    def add_item(self,e):
        if not self.current_file or not self.current_sheet: return
        dlg=ItemDialog(self,title="Tambah Task"); 
        if dlg.ShowModal()==wx.ID_OK:
            task,note,tanggal=dlg.get_values(); 
            if not task: return
            wb=openpyxl.load_workbook(self.current_file); ws=wb[self.current_sheet]; ws.append([task,0,note,tanggal]); wb.save(self.current_file); self.show_tasks()
        dlg.Destroy()
    def edit_item(self,task_name):
        wb=openpyxl.load_workbook(self.current_file); ws=wb[self.current_sheet]; row_idx=None; curr_note=""; curr_tanggal=""
        for i,row in enumerate(ws.iter_rows(min_row=2,values_only=True),start=2):
            t,s,n,d=safe_unpack(row)
            if str(t)==task_name: row_idx=i; curr_note=n or ""; curr_tanggal=d or ""; break
        if not row_idx: return
        dlg=ItemDialog(self,title="Edit Task",task=task_name,note=curr_note,tanggal=curr_tanggal)
        if dlg.ShowModal()==wx.ID_OK:
            new_task,new_note,new_tanggal=dlg.get_values(); ws.cell(row=row_idx,column=1).value=new_task; ws.cell(row=row_idx,column=3).value=new_note; ws.cell(row=row_idx,column=4).value=new_tanggal; wb.save(self.current_file); self.show_tasks()
        dlg.Destroy()
    def delete_item(self,task_name):
        wb=openpyxl.load_workbook(self.current_file); ws=wb[self.current_sheet]
        for i,row in enumerate(ws.iter_rows(min_row=2,values_only=True),start=2):
            t,_,_,_=safe_unpack(row)
            if str(t)==task_name: ws.delete_rows(i,1); break
        wb.save(self.current_file); self.show_tasks()
    def toggle_task(self,e,task_name):
        wb=openpyxl.load_workbook(self.current_file); ws=wb[self.current_sheet]
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value)==task_name: row[1].value=1 if e.IsChecked() else 0; break
        wb.save(self.current_file); self.show_tasks()


if __name__=="__main__":
    app=wx.App(); TodoApp(); app.MainLoop()
